"""
Excel Service for import/export
"""
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Optional
import os
from datetime import datetime
from models import Product, Sale
try:
    from .sale_service import SaleService
except ImportError:
    from sale_service import SaleService


class ExcelService:
    """Service for Excel import/export operations"""
    
    @staticmethod
    def _get_exports_dir():
        """Get exports directory path"""
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
        os.makedirs(exports_dir, exist_ok=True)
        return exports_dir
    
    @staticmethod
    def export_products(db: Session) -> str:
        """Export all products to Excel file"""
        products = db.query(Product).all()
        
        exports_dir = ExcelService._get_exports_dir()
        filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(exports_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Mahsulotlar"
        
        # Headers
        headers = [
            'ID', 'Nomi', 'Mahsulot kodi (Item Number)', 'Barcode', 'Brend', 'Yetkazib beruvchi', 'Joylashuv', '1 qop = dona',
            'Kelgan narx (dona)', 'Ulgurji narx (dona)', 'Dona narx (dona)', 'Oddiy narx (dona)',
            'Ulgurji qop narxi', 'Dona qop narxi', 'Oddiy qop narxi',
            'Ombordagi qop', 'Ombordagi dona', 'Jami dona', 'Rasm URL'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for product in products:
            # Calculate package prices (dona narxi * pieces_per_package)
            wholesale_package_price = (product.wholesale_price or 0.0) * (product.pieces_per_package or 1)
            retail_package_price = (product.retail_price or 0.0) * (product.pieces_per_package or 1)
            regular_package_price = (product.regular_price or 0.0) * (product.pieces_per_package or 1)
            
            ws.append([
                product.id,
                product.name,
                product.item_number or '',
                product.barcode or '',
                product.brand or '',
                product.supplier or '',
                product.location or '',
                product.pieces_per_package or 1,
                product.cost_price if product.cost_price is not None else 0.0,
                product.wholesale_price if product.wholesale_price is not None else 0.0,
                product.retail_price if product.retail_price is not None else 0.0,
                product.regular_price if product.regular_price is not None else 0.0,
                wholesale_package_price,
                retail_package_price,
                regular_package_price,
                product.packages_in_stock if product.packages_in_stock is not None else 0,
                product.pieces_in_stock if product.pieces_in_stock is not None else 0,
                product.total_pieces if product.total_pieces is not None else 0,
                product.image_url or ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def import_products(db: Session, file_path: str) -> dict:
        """Import products from Excel file"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Skip header row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:  # Skip empty rows (check name)
                continue
            
            try:
                # Parse row data (flexible column order)
                name = str(row[1]).strip() if row[1] else None
                if not name:
                    continue
                
                barcode = str(row[2]).strip() if len(row) > 2 and row[2] else None
                brand = str(row[3]).strip() if len(row) > 3 and row[3] else None
                supplier = str(row[4]).strip() if len(row) > 4 and row[4] else None
                location = str(row[5]).strip() if len(row) > 5 and row[5] else None
                pieces_per_package = int(row[6]) if len(row) > 6 and row[6] else 1
                cost_price = float(row[7]) if len(row) > 7 and row[7] else 0.0
                
                # New format: wholesale_price, retail_price, regular_price (all per piece)
                # Old format support: if package prices are provided, calculate piece prices
                wholesale_price = float(row[8]) if len(row) > 8 and row[8] else 0.0
                retail_price = float(row[9]) if len(row) > 9 and row[9] else 0.0
                regular_price = float(row[10]) if len(row) > 10 and row[10] else 0.0
                
                # If package prices are provided instead (old format), convert to piece prices
                if len(row) > 11 and row[11] and float(row[11]) > 0:
                    # Old format detected - package prices provided
                    wholesale_package_price = float(row[11]) if len(row) > 11 and row[11] else 0.0
                    retail_package_price = float(row[12]) if len(row) > 12 and row[12] else 0.0
                    regular_package_price = float(row[13]) if len(row) > 13 and row[13] else 0.0
                    packages_in_stock = int(row[14]) if len(row) > 14 and row[14] else 0
                    pieces_in_stock = int(row[15]) if len(row) > 15 and row[15] else 0
                    image_url = str(row[16]).strip() if len(row) > 16 and row[16] else None
                    
                    # Convert package prices to piece prices
                    if pieces_per_package > 0:
                        wholesale_price = wholesale_package_price / pieces_per_package
                        retail_price = retail_package_price / pieces_per_package
                        regular_price = regular_package_price / pieces_per_package
                else:
                    # New format - piece prices already provided
                    packages_in_stock = int(row[11]) if len(row) > 11 and row[11] else 0
                    pieces_in_stock = int(row[12]) if len(row) > 12 and row[12] else 0
                    image_url = str(row[13]).strip() if len(row) > 13 and row[13] else None
                
                # Check if product with same name or barcode exists
                existing = db.query(Product).filter(
                    (Product.name == name) | 
                    ((Product.barcode == barcode) if barcode else False)
                ).first()
                
                if existing:
                    errors.append(f"Qator {row_num}: '{name}' allaqachon mavjud (ID: {existing.id})")
                    continue
                
                product = Product(
                    name=name,
                    barcode=barcode if barcode and barcode.lower() != 'none' else None,
                    brand=brand if brand and brand.lower() != 'none' else None,
                    supplier=supplier if supplier and supplier.lower() != 'none' else None,
                    location=location if location and location.lower() != 'none' else None,
                    pieces_per_package=pieces_per_package,
                    cost_price=cost_price,
                    wholesale_price=wholesale_price,
                    retail_price=retail_price,
                    regular_price=regular_price,
                    packages_in_stock=packages_in_stock,
                    pieces_in_stock=pieces_in_stock,
                    image_url=image_url if image_url and image_url.lower() != 'none' else None
                )
                db.add(product)
                imported_count += 1
            except Exception as e:
                errors.append(f"Qator {row_num}: {str(e)}")
                continue
        
        db.commit()
        return {
            "imported": imported_count,
            "errors": errors
        }
    
    @staticmethod
    def export_sales(
        db: Session,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> str:
        """Export sales to Excel file"""
        sales = SaleService.get_sales(db)
        
        # Filter by date if provided
        if start_date or end_date:
            from datetime import datetime as dt
            filtered_sales = []
            for sale in sales:
                sale_date = sale.created_at.date()
                if start_date:
                    start = dt.fromisoformat(start_date).date()
                    if sale_date < start:
                        continue
                if end_date:
                    end = dt.fromisoformat(end_date).date()
                    if sale_date > end:
                        continue
                filtered_sales.append(sale)
            sales = filtered_sales
        
        exports_dir = ExcelService._get_exports_dir()
        filename = f"sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(exports_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sotuvlar"
        
        # Headers
        headers = [
            'ID', 'Sana', 'Sotuvchi', 'Mijoz', 'Summa'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for sale in sales:
            ws.append([
                sale.id,
                sale.created_at.strftime('%Y-%m-%d %H:%M'),
                sale.seller.name,
                sale.customer.name if sale.customer else "O'chirilgan mijoz",
                sale.total_amount
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def export_statistics(
        db: Session,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> str:
        """Export statistics to Excel file"""
        stats = SaleService.get_statistics(db, start_date, end_date)
        
        exports_dir = ExcelService._get_exports_dir()
        filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(exports_dir, filename)
        
        wb = Workbook()
        
        # Sheet 1: General Statistics
        ws1 = wb.active
        ws1.title = "Umumiy Statistika"
        
        # Headers
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws1.append(["Ko'rsatkich", "Qiymat"])
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # General stats
        ws1.append(["Jami sotuvlar", f"{stats.get('total_amount', 0):,.0f} so'm"])
        ws1.append(["Sotuvlar soni", f"{stats.get('total_sales', 0)} ta"])
        ws1.append(["O'rtacha sotuv", f"{stats.get('average_sale', 0):,.0f} so'm"])
        ws1.append(["Daromad", f"{stats.get('total_profit', 0):,.0f} so'm"])
        ws1.append(["Jami qarz", f"{stats.get('total_debt', 0):,.0f} so'm"])
        
        # Payment methods
        if 'payment_methods' in stats:
            ws1.append([])
            ws1.append(["TO'LOV TURLARI", ""])
            for cell in ws1[ws1.max_row]:
                cell.font = Font(bold=True)
            
            for method, data in stats['payment_methods'].items():
                method_name = {
                    'cash': '💵 Naqd',
                    'card': '💳 Karta',
                    'credit': '📝 Nasiya',
                    'bank_transfer': '🏦 O\'tkazma'
                }.get(method, method.upper())
                ws1.append([method_name, f"{data['amount']:,.0f} so'm ({data['count']} ta)"])
        
        # Sheet 2: Top Products
        if 'top_products' in stats:
            ws2 = wb.create_sheet("Top Mahsulotlar")
            ws2.append(["Mahsulot", "Mahsulot kodi (Item Number)", "Sotildi", "Summa"])
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for product in stats['top_products']:
                ws2.append([
                    product['name'],
                    product.get('item_number', '') or '',
                    product['quantity'],
                    f"{product['amount']:,.0f}"
                ])
        
        # Sheet 3: Top Customers
        if 'top_customers' in stats:
            ws3 = wb.create_sheet("Top Mijozlar")
            ws3.append(["Mijoz", "Xaridlar", "Summa"])
            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for customer in stats['top_customers']:
                ws3.append([customer['name'], customer['count'], f"{customer['amount']:,.0f}"])
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath

